from datetime import timedelta, date
from decimal import Decimal
from io import BytesIO
from typing import Tuple

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models.functions import TruncDate, TruncHour, TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook

from accounts.models import User, Role
from store.models import (
    Category,
    CategoryGroup,
    Product,
    OrderRequest,
    PickupPoint,
    Brand,
    Car,
)

from .forms import (
    CategoryForm,
    CategoryGroupForm,
    OrderRequestForm,
    UserForm,
    PickupPointForm,
    AutoPartProductForm,
    AutoGoodsProductForm,
    BrandForm,
    CarForm,
    SalesReportForm,
)

NAV_TABS = [
    ("overview", "Обзор"),
    ("categories", "Категории"),
    ("brands", "Бренды"),
    ("cars", "Автомобили"),
    ("products", "Товары"),
    ("orders", "Заказы"),
    ("users", "Пользователи"),
    ("pickups", "Пункты выдачи"),
    ("reports", "Отчёты"),
]
ADMIN_SECTIONS = {key for key, _ in NAV_TABS} | {"groups"}
MANAGER_SECTIONS = {"categories", "groups", "brands", "products", "orders"}


@login_required
def admin_panel(request):
    return _render_panel(request, mode="admin")


@login_required
def manager_panel(request):
    return _render_panel(request, mode="manager")


def _render_panel(request, mode: str):
    role_code = request.user.role_code
    if request.user.is_superuser:
        role_code = Role.RoleName.ADMIN
    elif role_code is None and mode == "admin" and request.user.is_staff:
        role_code = Role.RoleName.ADMIN
    allowed_roles = {Role.RoleName.ADMIN} if mode == "admin" else {Role.RoleName.MANAGER, Role.RoleName.ADMIN}
    if role_code not in allowed_roles:
        messages.error(request, "Недостаточно прав для доступа к этой панели.")
        return redirect("home")

    allowed_sections = ADMIN_SECTIONS if mode == "admin" else MANAGER_SECTIONS
    nav_tabs = [tab for tab in NAV_TABS if tab[0] in allowed_sections]
    default_section = nav_tabs[0][0] if nav_tabs else "overview"
    section = request.POST.get("section") or request.GET.get("section") or default_section
    if section not in allowed_sections:
        section = default_section
    cat_tab = request.POST.get("cat_tab") or request.GET.get("cat_tab") or "groups"
    product_tab = request.POST.get("product_tab") or request.GET.get("product_tab") or "auto_parts"
    analytics_range = request.GET.get("analytics_range") or "week"
    if analytics_range not in {"day", "week", "year"}:
        analytics_range = "week"
    report_tab = request.GET.get("report_tab") or "products"
    if report_tab not in {"products", "orders"}:
        report_tab = "products"

    report_initial = {
        "start_date": timezone.now().date() - timedelta(days=6),
        "end_date": timezone.now().date(),
    }
    report_form = SalesReportForm(initial=report_initial)
    report_rows = []
    report_total = Decimal("0")
    report_period_label = f"{report_initial['start_date'].strftime('%d.%m.%Y')} — {report_initial['end_date'].strftime('%d.%m.%Y')}"
    report_confirmed_only = False

    if section == "reports":
        if request.GET:
            report_form = SalesReportForm(request.GET)
        else:
            report_form = SalesReportForm(initial=report_initial)
        if report_form.is_valid():
            report_start = report_form.cleaned_data["start_date"]
            report_end = report_form.cleaned_data["end_date"]
        else:
            report_start = report_initial["start_date"]
            report_end = report_initial["end_date"]
        report_period_label = f"{report_start.strftime('%d.%m.%Y')} — {report_end.strftime('%d.%m.%Y')}"
        if report_tab == "orders":
            report_rows, report_total = _build_order_report_rows(report_start, report_end)
            if request.GET.get("export") == "1":
                return _export_orders_excel(report_rows, report_start, report_end, report_total)
        report_confirmed_only = report_form.cleaned_data.get("confirmed_only", False) if report_form.is_valid() else False
        if report_tab == "orders":
            report_rows, report_total = _build_order_report_rows(report_start, report_end, report_confirmed_only)
            if request.GET.get("export") == "1":
                return _export_orders_excel(report_rows, report_start, report_end, report_total, report_confirmed_only)
        else:
            report_rows, report_total = _build_sales_rows(report_start, report_end, report_confirmed_only)
            if request.GET.get("export") == "1":
                return _export_sales_excel(report_rows, report_start, report_end, report_total, report_confirmed_only)

    if request.method == "POST":
        handler = {
            "categories": _handle_category_post,
            "products": _handle_product_post,
            "orders": _handle_order_post,
            "users": _handle_user_post,
            "pickups": _handle_pickup_post,
            "groups": _handle_group_post,
            "brands": _handle_brand_post,
            "cars": _handle_car_post,
        }.get(section)
        if handler:
            response = handler(request)
            if response:
                return response

    analytics_range_options = [
        ("day", "День"),
        ("week", "Неделя"),
        ("year", "Год"),
    ]
    analytics_range_label = dict(analytics_range_options).get(analytics_range, "Неделя")

    confirmed_only = request.GET.get("confirmed_only") == "1"
    orders_qs = OrderRequest.objects.select_related("user", "pickup_point")
    if confirmed_only:
        orders_qs = orders_qs.filter(status="confirmed")
    orders_list = list(orders_qs.order_by("-created_at")[:50])

    context = {
        "section": section,
        "stats": _collect_stats(),
        "order_status_stats": _collect_order_status_stats(),
        "orders_period_counts": _collect_order_period_counts(),
        "orders_chart": _collect_analytics_series(OrderRequest.objects.all(), "created_at", analytics_range),
        "users_chart": _collect_analytics_series(User.objects.all(), "date_joined", analytics_range),
        "category_form": CategoryForm(),
        "categories": Category.objects.select_related("group").order_by("main_category", "name"),
        "group_form": CategoryGroupForm(),
        "groups": CategoryGroup.objects.order_by("main_category", "order"),
        "auto_part_form": AutoPartProductForm(),
        "auto_goods_form": AutoGoodsProductForm(),
        "auto_parts": Product.objects.select_related("category", "brand", "auto_part_spec")
        .filter(product_type=Product.ProductType.AUTO_PART)
        .order_by("-id")[:50],
        "auto_goods": Product.objects.select_related("category", "brand", "auto_goods_spec")
        .filter(product_type=Product.ProductType.AUTO_GOODS)
        .order_by("-id")[:50],
        "order_form": OrderRequestForm(),
        "orders": orders_list,
        "user_form": UserForm(),
        "users": User.objects.order_by("-date_joined")[:50],
        "pickup_form": PickupPointForm(),
        "pickup_points": PickupPoint.objects.order_by("-created_at"),
        "brands": Brand.objects.order_by("name"),
        "brand_form": BrandForm(),
        "car_form": CarForm(),
        "cars": Car.objects.order_by("make", "model", "generation"),
        "delivery_choices": OrderRequest.DELIVERIES,
        "payment_choices": OrderRequest.PAYMENTS,
        "nav_tabs": nav_tabs,
        "main_category_choices": Category.MainCategory.choices,
        "cat_tab": cat_tab,
        "product_tab": product_tab,
        "report_tab": report_tab,
        "report_tab_options": [("products", "По товарам"), ("orders", "По заказам")],
        "panel_mode": mode,
        "panel_title": "Панель администратора" if mode == "admin" else "Панель менеджера",
        "role_choices": Role.objects.filter(role__in=Role.RoleName.values).order_by("role"),
        "analytics_range": analytics_range,
        "analytics_range_options": analytics_range_options,
        "analytics_range_label": analytics_range_label,
        "report_form": report_form,
        "report_rows": report_rows,
        "report_total": report_total,
        "report_period_label": report_period_label,
        "orders_confirmed_only": confirmed_only,
        "report_confirmed_only": report_confirmed_only,
    }
    return render(request, "dashboard/admin_panel.html", context)


def _collect_stats():
    return {
        "categories": Category.objects.count(),
        "products": Product.objects.count(),
        "orders": OrderRequest.objects.count(),
        "users": User.objects.count(),
    }


def _collect_order_status_stats():
    base = {"new": 0, "confirmed": 0, "declined": 0}
    totals = (
        OrderRequest.objects.values("status")
        .annotate(total=Count("pk"))
    )
    for row in totals:
        status = row.get("status") or "new"
        base[status] = row["total"]
    return base


def _collect_order_period_counts():
    now = timezone.now()
    start_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_month = start_day.replace(day=1)
    start_year = start_day.replace(month=1, day=1)
    return {
        "day": OrderRequest.objects.filter(created_at__gte=start_day).count(),
        "month": OrderRequest.objects.filter(created_at__gte=start_month).count(),
        "year": OrderRequest.objects.filter(created_at__gte=start_year).count(),
    }


def _collect_analytics_series(queryset, date_field: str, period: str):
    period = period if period in {"day", "week", "year"} else "week"
    now = timezone.now()

    if period == "day":
        start = (now - timedelta(hours=23)).replace(minute=0, second=0, microsecond=0)
        buckets = []
        for i in range(24):
            point = start + timedelta(hours=i)
            buckets.append((point, point.strftime("%H:%M")))
        trunc_expr = TruncHour(date_field)
        filter_kwargs = {f"{date_field}__gte": start}
        bucket_key_cast = lambda value: value
    elif period == "year":
        today = now.date()
        first_day_current = today.replace(day=1)
        year = first_day_current.year
        month = first_day_current.month - 11
        while month <= 0:
            month += 12
            year -= 1
        buckets = []
        current_year = year
        current_month = month
        for _ in range(12):
            bucket_date = date(current_year, current_month, 1)
            buckets.append((bucket_date, bucket_date.strftime("%b")))
            current_year, current_month = _increment_month(current_year, current_month)
        trunc_expr = TruncMonth(date_field)
        filter_kwargs = {f"{date_field}__date__gte": buckets[0][0]}
        bucket_key_cast = lambda value: value.date().replace(day=1)
    else:  # week
        start = (now - timedelta(days=6)).date()
        buckets = []
        for i in range(7):
            bucket_date = start + timedelta(days=i)
            buckets.append((bucket_date, bucket_date.strftime("%d.%m")))
        trunc_expr = TruncDate(date_field)
        filter_kwargs = {f"{date_field}__date__gte": start}
        bucket_key_cast = lambda value: value

    totals = {
        bucket_key_cast(row["bucket"]): row["total"]
        for row in queryset.filter(**filter_kwargs)
        .annotate(bucket=trunc_expr)
        .values("bucket")
        .annotate(total=Count("pk"))
    }

    data = []
    max_value = 0
    for bucket_value, label in buckets:
        count = totals.get(bucket_value, 0)
        max_value = max(max_value, count)
        data.append({"label": label, "count": count})
    max_value = max_value or 1
    for entry in data:
        entry["percent"] = (entry["count"] / max_value) * 100 if max_value else 0
    return {"data": data, "max": max_value, "period": period}


def _increment_month(year: int, month: int) -> Tuple[int, int]:
    month += 1
    if month > 12:
        month = 1
        year += 1
    return year, month


def _build_sales_rows(start_date, end_date, confirmed_only=False):
    rows = []
    total = Decimal("0")
    qs = (
        OrderRequest.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        .order_by("created_at")
    )
    if confirmed_only:
        qs = qs.filter(status="confirmed")
    for order in qs:
        snapshot = order.cart_snapshot or []
        for item in snapshot:
            quantity = int(item.get("quantity") or 0)
            price = Decimal(str(item.get("price") or 0))
            line_total = Decimal(str(item.get("line_total") or (price * quantity)))
            rows.append(
                {
                    "order_id": order.id,
                    "date": order.created_at,
                    "product": item.get("name") or f"Товар #{item.get('product_id')}",
                    "quantity": quantity,
                    "price": price,
                    "line_total": line_total,
                }
            )
            total += line_total
    return rows, total


def _export_sales_excel(rows, start_date, end_date, total_amount, confirmed_only=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    ws["A1"] = "Отчёт по продажам"
    note = f"Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
    if confirmed_only:
        note += " (только подтверждённые)"
    ws["A2"] = note
    ws.append([])
    headers = ["ID заказа", "Дата", "Товар", "Кол-во", "Цена", "Сумма"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["order_id"],
                row["date"].strftime("%d.%m.%Y %H:%M"),
                row["product"],
                row["quantity"],
                float(row["price"]),
                float(row["line_total"]),
            ]
        )
    ws.append([])
    ws.append(["Итого", "", "", "", "", float(total_amount)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"sales_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_order_report_rows(start_date, end_date, confirmed_only=False):
    rows = []
    total = Decimal("0")
    qs = (
        OrderRequest.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        .select_related("user")
        .order_by("created_at")
    )
    if confirmed_only:
        qs = qs.filter(status="confirmed")
    for order in qs:
        total += order.total_amount
        rows.append(
            {
                "order_id": order.id,
                "date": order.created_at,
                "customer": order.full_name or order.email or (order.user.email if order.user else "—"),
                "total": order.total_amount,
                "status": order.status,
            }
        )
    return rows, total


def _export_orders_excel(rows, start_date, end_date, total_amount, confirmed_only=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    ws["A1"] = "Отчёт по заказам"
    note = f"Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
    if confirmed_only:
        note += " (только подтверждённые)"
    ws["A2"] = note
    ws.append([])
    headers = ["ID заказа", "Дата", "Клиент", "Сумма", "Статус"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["order_id"],
                row["date"].strftime("%d.%m.%Y %H:%M"),
                row["customer"],
                float(row["total"]),
                row["status"],
            ]
        )
    ws.append([])
    ws.append(["Итого", "", "", float(total_amount), ""])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"orders_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _handle_category_post(request):
    action = request.POST.get("action")
    if action == "delete":
        category = get_object_or_404(Category, pk=request.POST.get("object_id"))
        category.delete()
        messages.success(request, "Подкатегория удалена.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(Category, pk=request.POST.get("object_id"))
        form = CategoryForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Категория сохранена.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    cat_tab = request.POST.get("cat_tab", "subcategories")
    return redirect(f"{request.path}?section=categories&cat_tab={cat_tab}")


def _handle_group_post(request):
    action = request.POST.get("action")
    if action == "delete":
        group = get_object_or_404(CategoryGroup, pk=request.POST.get("object_id"))
        group.delete()
        messages.success(request, "Группа удалена.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(CategoryGroup, pk=request.POST.get("object_id"))
        form = CategoryGroupForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Группа сохранена.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=categories&cat_tab=groups")


def _handle_product_post(request):
    action = request.POST.get("action")
    tab = request.POST.get("product_tab", "auto_parts")
    if action == "delete":
        product = get_object_or_404(Product, pk=request.POST.get("object_id"))
        tab = (
            "auto_parts"
            if product.product_type == Product.ProductType.AUTO_PART
            else "auto_goods"
        )
        product.delete()
        messages.success(request, "Товар удалён.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(Product, pk=request.POST.get("object_id"))
        if tab == "auto_goods":
            form = AutoGoodsProductForm(request.POST, request.FILES, instance=instance)
        else:
            form = AutoPartProductForm(request.POST, request.FILES, instance=instance)
            tab = "auto_parts"
        if form.is_valid():
            form.save()
            messages.success(request, "Товар сохранён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=products&product_tab={tab}")


def _handle_order_post(request):
    instance = get_object_or_404(OrderRequest, pk=request.POST.get("object_id"))
    action = request.POST.get("action", "update")

    if action == "confirm":
        instance.status = "confirmed"
        instance.save(update_fields=["status"])
        messages.success(request, "Заказ подтверждён.")
    elif action == "decline":
        instance.status = "declined"
        instance.save(update_fields=["status"])
        messages.info(request, "Заказ отменён.")
    else:
        form = OrderRequestForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Заказ обновлён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=orders")


def _handle_user_post(request):
    action = request.POST.get("action")
    if action == "delete":
        user_id = request.POST.get("object_id")
        if not user_id:
            messages.error(request, "Не удалось определить пользователя для удаления.")
            return redirect(f"{request.path}?section=users")
        user = get_object_or_404(User, pk=user_id)
        if user.is_superuser:
            messages.error(request, "Нельзя удалить суперпользователя.")
        else:
            user.delete()
            messages.success(request, "Пользователь удалён.")
    elif action == "set_role":
        user_id = request.POST.get("object_id")
        role_id = request.POST.get("role_id")
        if not user_id or not role_id:
            messages.error(request, "Выберите пользователя и роль перед сохранением.")
            return redirect(f"{request.path}?section=users")
        user = get_object_or_404(User, pk=user_id)
        role = get_object_or_404(Role, pk=role_id)
        user.role = role
        user.is_staff = role.role == Role.RoleName.ADMIN
        user.save(update_fields=["role", "is_staff"])
        messages.success(request, "Роль пользователя обновлена.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(User, pk=request.POST.get("object_id"))
        form = UserForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Пользователь сохранён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=users")


def _handle_pickup_post(request):
    action = request.POST.get("action")
    if action == "delete":
        pvz = get_object_or_404(PickupPoint, pk=request.POST.get("object_id"))
        pvz.delete()
        messages.success(request, "ПВЗ удалён.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(PickupPoint, pk=request.POST.get("object_id"))
        form = PickupPointForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "ПВЗ сохранён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=pickups")


def _handle_brand_post(request):
    action = request.POST.get("action")
    if action == "delete":
        brand_id = request.POST.get("object_id")
        if not brand_id:
            messages.error(request, "Не удалось определить бренд. Попробуйте ещё раз.")
            return redirect(f"{request.path}?section=brands")
        brand = get_object_or_404(Brand, pk=brand_id)
        # Prevent deleting if products exist?
        if brand.product_set.exists():
            messages.error(request, "Удалите или переназначьте товары этого бренда перед удалением.")
        else:
            brand.delete()
            messages.success(request, "Бренд удалён.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(Brand, pk=request.POST.get("object_id"))
        form = BrandForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Бренд сохранён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=brands")


def _handle_car_post(request):
    action = request.POST.get("action")
    if action == "delete":
        car = get_object_or_404(Car, pk=request.POST.get("object_id"))
        car.delete()
        messages.success(request, "Автомобиль удалён.")
    else:
        instance = None
        if action == "update":
            instance = get_object_or_404(Car, pk=request.POST.get("object_id"))
        form = CarForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Автомобиль сохранён.")
        else:
            messages.error(request, f"Ошибка: {form.errors}")
    return redirect(f"{request.path}?section=cars")


@login_required
def orders_feed(request):
    role_code = getattr(request.user, "role_code", None)
    if request.user.is_superuser or request.user.is_staff:
        role_code = Role.RoleName.ADMIN
    allowed = {Role.RoleName.ADMIN, Role.RoleName.MANAGER}
    if role_code not in allowed:
        return JsonResponse({"error": "forbidden"}, status=403)
    last_id = request.GET.get("last_id")
    try:
        last_id = int(last_id or 0)
    except (TypeError, ValueError):
        last_id = 0
    new_orders = (
        OrderRequest.objects.select_related("pickup_point")
        .filter(id__gt=last_id)
        .order_by("-created_at")[:10]
    )
    html = render_to_string(
        "dashboard/partials/order_cards.html",
        {"orders": new_orders, "show_empty": False, "highlight_new": True},
        request=request,
    )
    max_id = last_id
    if new_orders:
        max_id = max(order.id for order in new_orders)
    message = ""
    if new_orders:
        newest = new_orders.first()
        message = f"Новый заказ #{newest.id}"
    return JsonResponse(
        {
            "html": html,
            "has_new": bool(new_orders),
            "last_id": max_id,
            "message": message,
        }
    )
